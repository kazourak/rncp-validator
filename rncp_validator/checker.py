"""Checker to identify commits not done on school time."""

import argparse
import itertools
import os
import shutil

from openpyxl import load_workbook

from rncp_validator.Calendar import Calendar
from rncp_validator.tools import get_all_commits, get_calendars


def analyse(calendar_path: str, git_path: str, branch: str = None):
    """
    Check for if the given git history match with the calendar.
    :param calendar_path: The path to the calendar as a xlsx file.
    :param git_path: The .git path.
    :param branch: The branch to check.
    :return: None
    """

    print("---------------------------------------")
    print("\033[1m" + f"Working on {calendar_path} with {git_path}" + "\033[0m")

    cl = Calendar(load_workbook(calendar_path))
    cl.get_periods()

    commit_list = get_all_commits(git_path, branch=branch)

    print("---------------------------------------")

    for commit, (date, author) in commit_list.items():
        if cl.date_in_period(date):
            print(
                "\033[92m" + "Commit",
                commit,
                "by",
                author,
                "at",
                date,
                "is in a school period" + "\033[0m",
            )
        else:
            print(
                "\033[91m" + "Commit",
                commit,
                "by",
                author,
                "at",
                date,
                "is not in a school period" + "\033[0m",
            )


def get_bad_commits(calendar_path: str, git_path: str, branch: str = None) -> dict:
    """
    Returns all commits that don't match the given calendar.
    :param calendar_path: The path to the calendar as a xlsx file.
    :param git_path: The .git path.
    :param branch: The branch to check.
    :return: All bad commits as a dict.
    """

    bad_commits = {}

    cl = Calendar(load_workbook(calendar_path))
    cl.get_periods()

    commit_list = get_all_commits(git_path, branch=branch)

    for commit, (date, author) in commit_list.items():
        if not cl.date_in_period(date):
            bad_commits[commit] = (date, author)

    return bad_commits


def main():
    """
    Main program.
    """
    try:
        parser = argparse.ArgumentParser(description="Process calendar and git parse arguments.")
        parser.add_argument(
            "calendar_path", type=str, help="Path to the calendar file or directory"
        )
        parser.add_argument("git_paths", type=str, nargs="+", help="Git paths arguments")
        parser.add_argument("--branch", type=str, default=None, help="Branch to check")
        parser.add_argument(
            "--not_recursive",
            action="store_false",
            help="Do not explore the directory recursively.",
        )

        args = parser.parse_args()

        xlsx_paths = get_calendars(args.calendar_path, args.not_recursive)

        for xlsx_path, git_path in itertools.product(xlsx_paths, args.git_paths):
            analyse(xlsx_path, git_path, args.branch)

        if os.path.exists("/tmp/rncp-validator"):
            shutil.rmtree("/tmp/rncp-validator")

    except Exception as e:
        raise ValueError("{Find an error}") from e


if __name__ == "__main__":
    main()
