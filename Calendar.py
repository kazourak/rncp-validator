from SchoolPeriod import SchoolPeriod
from openpyxl import load_workbook
from datetime import datetime, timedelta


def month_to_number(month_name: str):
    month = {
        "Janvier": "01",
        "Février": "02",
        "Mars": "03",
        "Avril": "04",
        "Mai": "05",
        "Juin": "06",
        "Juillet": "07",
        "Août": "08",
        "Septembre": "09",
        "Octobre": "10",
        "Novembre": "11",
        "Décembre": "12"
    }

    return month[month_name]


def to_date(month_str: str, day_str: str, year_str: str):
    return datetime(int(year_str), int(month_str), int(day_str))


class Calendar():
    def __init__(self, calendar_file):
        self.calendar = calendar_file
        self.periods = []

    def get_periods(self):
        for sheet in self.calendar.sheetnames:
            for column in self.calendar[sheet].iter_cols(min_row=6, max_row=38, min_col=1, max_col=12):
                periods = []
                current_period = []
                prev_date = None

                for cell in column:
                    if cell.fill.bgColor.rgb == "0061b3ff" or cell.fill.bgColor.rgb == "002cf28f":
                        current_date = to_date(month_to_number(column[0].value), cell.value.split()[1], sheet)

                        if not current_period:
                            current_period.append(current_date)
                        elif prev_date and current_date == prev_date + timedelta(days=1):
                            current_period.append(current_date)
                        else:
                            periods.append((current_period[0], current_period[-1]))
                            current_period = [current_date]

                        prev_date = current_date

                if current_period:
                    periods.append((current_period[0], current_period[-1]))

                for start_date, end_date in periods:
                    self.periods.append(SchoolPeriod(start_date, end_date))

    def date_in_period(self, date_to_compare):
        if not 8 <= date_to_compare.hour < 20:
            return False
        for period in self.periods:
            if period.in_date_range(date_to_compare):
                return True
        return False
